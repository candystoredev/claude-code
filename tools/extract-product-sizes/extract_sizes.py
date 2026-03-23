#!/usr/bin/env python3
"""
Extract OZ size data from product names in an Excel file.

Usage:
    python extract_sizes.py input.xlsx
    python extract_sizes.py input.xlsx --column "Product Name"
    python extract_sizes.py input.xlsx --column "Product Name" --output output.xlsx

The script adds a "Size" column next to the product name column containing
the extracted size (e.g. "0.74 OZ", "3.5OZ").
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def extract_size(name: str) -> str:
    """Extract the OZ size substring from a product name.

    Handles both '0.74 OZ' (space) and '3.5OZ' (no space) formats.
    Returns the matched substring as-is, e.g. '0.74 OZ' or '3.5OZ'.
    Returns empty string if no match found.
    """
    if not name:
        return ""
    match = re.search(r"(\d*\.?\d+)\s*OZ", str(name), re.IGNORECASE)
    return match.group(0) if match else ""


def find_column(ws, column_name: str) -> int | None:
    """Return the 1-based column index matching column_name in the header row."""
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() == column_name.strip().lower():
            return col_idx
    return None


def process_file(input_path: Path, column_name: str, output_path: Path) -> None:
    print(f"Loading {input_path} ...")
    wb = load_workbook(input_path)
    ws = wb.active

    # Locate the product name column
    col_idx = find_column(ws, column_name)
    if col_idx is None:
        # Fall back to column A if the named column isn't found
        print(
            f"Warning: column '{column_name}' not found in header row. "
            "Falling back to column A."
        )
        col_idx = 1

    # Insert a new "Size" column immediately after the product name column
    size_col_idx = col_idx + 1
    ws.insert_cols(size_col_idx)
    ws.cell(row=1, column=size_col_idx).value = "Size"

    total_rows = ws.max_row - 1  # exclude header
    matched = 0

    for row in range(2, ws.max_row + 1):
        product_name = ws.cell(row=row, column=col_idx).value
        size = extract_size(product_name or "")
        ws.cell(row=row, column=size_col_idx).value = size
        if size:
            matched += 1

    print(f"Processed {total_rows} rows — {matched} sizes extracted.")
    wb.save(output_path)
    print(f"Saved to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract OZ sizes from product names in Excel.")
    parser.add_argument("input", type=Path, help="Input Excel file (.xlsx)")
    parser.add_argument(
        "--column",
        default="Product Name",
        help="Header name of the product name column (default: 'Product Name')",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output file path (default: input file with '_with_sizes' suffix)",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Error: file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    output_path = args.output or args.input.with_stem(args.input.stem + "_with_sizes")
    process_file(args.input, args.column, output_path)


if __name__ == "__main__":
    main()
