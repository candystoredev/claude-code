import csv
import io
import json
import mimetypes
import os
import re
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

import boto3
import requests
from botocore.exceptions import BotoCoreError, ClientError
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, render_template, request, send_file, session
from openpyxl import Workbook, load_workbook

load_dotenv()

VERSION = "1.1.2"

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB upload limit

APP_PASSWORD = os.environ.get("APP_PASSWORD", "")

# Cloud storage pre-configuration (optional — set in Railway env vars)
STORAGE_TYPE = os.environ.get("STORAGE_TYPE", "r2")  # "r2" or "s3"

R2_ACCOUNT_ID      = os.environ.get("R2_ACCOUNT_ID", "")
R2_BUCKET_NAME     = os.environ.get("R2_BUCKET_NAME", "")
R2_ACCESS_KEY_ID   = os.environ.get("R2_ACCESS_KEY_ID", "")
R2_SECRET_KEY      = os.environ.get("R2_SECRET_KEY", "")
R2_PUBLIC_BASE_URL = os.environ.get("R2_PUBLIC_BASE_URL", "")
R2_FOLDER_PREFIX   = os.environ.get("R2_FOLDER_PREFIX", "")

S3_BUCKET_NAME     = os.environ.get("S3_BUCKET_NAME", "")
S3_ACCESS_KEY_ID   = os.environ.get("S3_ACCESS_KEY_ID", "")
S3_SECRET_KEY      = os.environ.get("S3_SECRET_KEY", "")
S3_PUBLIC_BASE_URL = os.environ.get("S3_PUBLIC_BASE_URL", "")
S3_FOLDER_PREFIX   = os.environ.get("S3_FOLDER_PREFIX", "")
S3_REGION          = os.environ.get("S3_REGION", "us-east-1")

R2_PRECONFIGURED = all([R2_ACCOUNT_ID, R2_BUCKET_NAME, R2_ACCESS_KEY_ID, R2_SECRET_KEY, R2_PUBLIC_BASE_URL])
S3_PRECONFIGURED = all([S3_BUCKET_NAME, S3_ACCESS_KEY_ID, S3_SECRET_KEY, S3_PUBLIC_BASE_URL])

# In-memory job store: job_id -> {"rows": [...], "events": [...], "done": bool, "output_bytes": bytes|None}
_jobs: dict = {}
_jobs_lock = threading.Lock()

MAX_WORKERS = 8
JOB_TTL_SECONDS = 3600  # 1 hour
MAX_JOBS = 500           # evict oldest when exceeded


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def _is_authenticated() -> bool:
    if not APP_PASSWORD:
        return True
    return session.get("authenticated") is True


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    if not _is_authenticated():
        return render_template("login.html")
    return render_template("index.html",
        version=VERSION,
        storage_type=STORAGE_TYPE,
        r2=dict(account_id=R2_ACCOUNT_ID, bucket_name=R2_BUCKET_NAME,
                access_key_id=R2_ACCESS_KEY_ID,
                secret_key="" if R2_PRECONFIGURED else R2_SECRET_KEY,
                public_base_url=R2_PUBLIC_BASE_URL, folder_prefix=R2_FOLDER_PREFIX,
                preconfigured=R2_PRECONFIGURED),
        s3=dict(bucket_name=S3_BUCKET_NAME, access_key_id=S3_ACCESS_KEY_ID,
                secret_key="" if S3_PRECONFIGURED else S3_SECRET_KEY,
                public_base_url=S3_PUBLIC_BASE_URL,
                folder_prefix=S3_FOLDER_PREFIX, region=S3_REGION,
                preconfigured=S3_PRECONFIGURED),
    )


@app.route("/login", methods=["POST"])
def login():
    pw = request.form.get("password", "")
    if pw == APP_PASSWORD:
        session["authenticated"] = True
        return Response("", status=302, headers={"Location": "/"})
    return render_template("login.html", error="Incorrect password.")


@app.route("/logout")
def logout():
    session.clear()
    return Response("", status=302, headers={"Location": "/"})


@app.route("/process", methods=["POST"])
def process():
    if not _is_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    # --- Parse form fields (fall back to env vars when pre-configured) ---
    storage_type = request.form.get("storage_type", "") or STORAGE_TYPE
    if storage_type == "r2":
        access_key_id     = request.form.get("access_key_id", "").strip()     or R2_ACCESS_KEY_ID
        secret_access_key = request.form.get("secret_access_key", "").strip() or R2_SECRET_KEY
        bucket_name       = request.form.get("bucket_name", "").strip()       or R2_BUCKET_NAME
        public_base_url   = (request.form.get("public_base_url", "").strip()  or R2_PUBLIC_BASE_URL).rstrip("/")
        folder_prefix     = (request.form.get("folder_prefix", "").strip()    or R2_FOLDER_PREFIX).strip("/")
        account_id        = request.form.get("account_id", "").strip()        or R2_ACCOUNT_ID
        aws_region        = "auto"
    else:
        access_key_id     = request.form.get("access_key_id", "").strip()     or S3_ACCESS_KEY_ID
        secret_access_key = request.form.get("secret_access_key", "").strip() or S3_SECRET_KEY
        bucket_name       = request.form.get("bucket_name", "").strip()       or S3_BUCKET_NAME
        public_base_url   = (request.form.get("public_base_url", "").strip()  or S3_PUBLIC_BASE_URL).rstrip("/")
        folder_prefix     = (request.form.get("folder_prefix", "").strip()    or S3_FOLDER_PREFIX).strip("/")
        account_id        = ""
        aws_region        = request.form.get("aws_region", "").strip()        or S3_REGION

    if not all([access_key_id, secret_access_key, bucket_name, public_base_url]):
        return jsonify({"error": "Missing required fields: access_key_id, secret_access_key, bucket_name, public_base_url"}), 400

    if storage_type == "r2" and not account_id:
        return jsonify({"error": "Account ID is required for Cloudflare R2"}), 400

    # --- Parse uploaded file ---
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    filename = f.filename or ""

    try:
        rows, url_col, name_col = _parse_spreadsheet(f, filename)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not rows:
        return jsonify({"error": "No rows found in the uploaded file"}), 400

    # --- Build S3 client config ---
    if storage_type == "r2":
        endpoint_url = f"https://{account_id}.r2.cloudflarestorage.com"
        region_name = "auto"
    else:
        endpoint_url = None
        region_name = aws_region

    s3_config = {
        "storage_type": storage_type,
        "access_key_id": access_key_id,
        "secret_access_key": secret_access_key,
        "bucket_name": bucket_name,
        "endpoint_url": endpoint_url,
        "region_name": region_name,
        "public_base_url": public_base_url,
        "folder_prefix": folder_prefix,
    }

    # --- Evict oldest jobs if over limit ---
    with _jobs_lock:
        if len(_jobs) > MAX_JOBS:
            sorted_ids = sorted(_jobs.keys(), key=lambda jid: _jobs[jid].get("created_at", 0))
            for jid in sorted_ids[:100]:
                del _jobs[jid]

    # --- Create job and start background thread ---
    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "rows": rows,
            "total": len(rows),
            "events": [],
            "done": False,
            "output_bytes": None,
            "created_at": time.time(),
        }

    thread = threading.Thread(target=_run_job, args=(job_id, rows, s3_config), daemon=True)
    thread.start()

    return jsonify({"job_id": job_id, "total": len(rows), "url_col": url_col, "name_col": name_col})


@app.route("/stream/<job_id>")
def stream(job_id: str):
    if not _is_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    def generate():
        sent = 0
        while True:
            with _jobs_lock:
                job = _jobs.get(job_id)
                if job is None:
                    yield f"data: {json.dumps({'error': 'Job not found'})}\n\n"
                    return
                events_snapshot = list(job["events"])
                is_done = job["done"]

            while sent < len(events_snapshot):
                yield f"data: {json.dumps(events_snapshot[sent])}\n\n"
                sent += 1

            if is_done and sent >= len(events_snapshot):
                return

            time.sleep(0.2)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/download/<job_id>")
def download(job_id: str):
    if not _is_authenticated():
        return jsonify({"error": "Unauthorized"}), 401

    with _jobs_lock:
        job = _jobs.get(job_id)

    if not job:
        return jsonify({"error": "Results not found or expired — please re-run the batch"}), 404

    if not job["done"]:
        return jsonify({"error": "Job not ready"}), 404

    output_bytes = job.get("output_bytes")
    if not output_bytes:
        return jsonify({"error": "No output available"}), 500

    buf = io.BytesIO(output_bytes)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="image-batch-renamer-results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Background job
# ---------------------------------------------------------------------------

def _run_job(job_id: str, rows: list[dict], s3_config: dict):
    results = []
    seen_keys: set[str] = set()
    seen_keys_lock = threading.Lock()

    def process_one(idx_row):
        idx, row = idx_row
        original_url = row.get("original_url", "").strip()
        new_filename = row.get("new_filename", "").strip()

        if not original_url:
            return {
                "index": idx, "total": len(rows),
                "original_url": original_url, "new_filename": "",
                "new_url": "", "status": "error", "message": "Empty URL — skipped",
            }

        # Validate URL scheme (prevent SSRF)
        parsed = urlparse(original_url)
        if parsed.scheme not in ("http", "https"):
            return {
                "index": idx, "total": len(rows),
                "original_url": original_url, "new_filename": "",
                "new_url": "", "status": "error",
                "message": f"Invalid URL scheme '{parsed.scheme}' — only http/https allowed",
            }

        try:
            # 1. Download image
            resp = requests.get(original_url, timeout=30, stream=True,
                                headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            image_bytes = resp.content
            content_type = resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()

            # 2. Determine final filename
            final_filename = _resolve_filename(new_filename, original_url, content_type)

            # 3. Build S3 key, resolving collisions
            prefix = s3_config["folder_prefix"]
            base_key = f"{prefix}/{final_filename}" if prefix else final_filename

            with seen_keys_lock:
                key = base_key
                if key in seen_keys:
                    stem, ext = os.path.splitext(final_filename)
                    counter = 2
                    while key in seen_keys:
                        suffixed = f"{stem}-{counter}{ext}"
                        key = f"{prefix}/{suffixed}" if prefix else suffixed
                        counter += 1
                    final_filename = os.path.basename(key)
                seen_keys.add(key)

            collision_note = f" (renamed: collision with earlier row)" if key != base_key else ""

            # 4. Upload to S3/R2
            client = boto3.client(
                "s3",
                endpoint_url=s3_config["endpoint_url"],
                region_name=s3_config["region_name"],
                aws_access_key_id=s3_config["access_key_id"],
                aws_secret_access_key=s3_config["secret_access_key"],
            )

            put_kwargs = {
                "Bucket": s3_config["bucket_name"],
                "Key": key,
                "Body": image_bytes,
                "ContentType": content_type,
            }
            # ACL not set — both S3 and R2 use bucket-level public access policy

            client.put_object(**put_kwargs)

            # 5. Construct new URL
            new_url = f"{s3_config['public_base_url']}/{key}"

            return {
                "index": idx, "total": len(rows),
                "original_url": original_url, "new_filename": final_filename,
                "new_url": new_url, "status": "success", "message": collision_note.strip(),
            }

        except requests.RequestException as e:
            return {
                "index": idx, "total": len(rows),
                "original_url": original_url, "new_filename": new_filename,
                "new_url": "", "status": "error", "message": f"Download failed: {e}",
            }
        except (BotoCoreError, ClientError) as e:
            return {
                "index": idx, "total": len(rows),
                "original_url": original_url, "new_filename": new_filename,
                "new_url": "", "status": "error",
                "message": (
                    f"Upload failed: {e} "
                    f"[bucket={s3_config['bucket_name']}, "
                    f"region={s3_config['region_name']}, "
                    f"endpoint={s3_config['endpoint_url']}, "
                    f"key={key}]"
                ),
            }
        except Exception as e:
            return {
                "index": idx, "total": len(rows),
                "original_url": original_url, "new_filename": new_filename,
                "new_url": "", "status": "error", "message": f"Unexpected error: {e}",
            }

    try:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_one, (i, row)): i for i, row in enumerate(rows)}
            for future in as_completed(futures):
                try:
                    result = future.result()
                except Exception as e:
                    idx = futures[future]
                    result = {
                        "index": idx, "total": len(rows),
                        "original_url": rows[idx].get("original_url", ""),
                        "new_filename": "", "new_url": "",
                        "status": "error", "message": f"Internal error: {e}",
                    }
                results.append(result)
                with _jobs_lock:
                    _jobs[job_id]["events"].append(result)
    finally:
        # Always mark job done, even if something went wrong
        results.sort(key=lambda r: r["index"])
        output_bytes = _build_output_excel(results)
        success_count = sum(1 for r in results if r["status"] == "success")
        with _jobs_lock:
            _jobs[job_id]["output_bytes"] = output_bytes
            _jobs[job_id]["done"] = True
            _jobs[job_id]["events"].append({
                "type": "done",
                "total": len(rows),
                "success_count": success_count,
                "error_count": len(rows) - success_count,
            })

        # Schedule TTL cleanup
        def _ttl_cleanup():
            time.sleep(JOB_TTL_SECONDS)
            with _jobs_lock:
                _jobs.pop(job_id, None)
        threading.Thread(target=_ttl_cleanup, daemon=True).start()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_spreadsheet(file_obj, filename: str) -> tuple[list[dict], str, str]:
    """Parse CSV or Excel into list of dicts with 'original_url' and 'new_filename'.
    Returns (rows, url_col_name, name_col_name).
    """
    name = filename.lower()

    if name.endswith(".csv"):
        content = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        raw_rows = list(reader)
        headers = list(reader.fieldnames or [])
    elif name.endswith((".xlsx", ".xls")):
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], "", ""
            headers = [str(h or "").strip() for h in header_row]
            raw_rows = [dict(zip(headers, row)) for row in rows_iter]
        finally:
            wb.close()
    else:
        raise ValueError("Unsupported file format. Upload a .csv, .xlsx, or .xls file.")

    url_col = _find_column(headers, ["original_url", "url", "image_url", "image src", "src"])
    name_col = _find_column(headers, ["new_filename", "filename", "new_name", "new name"])

    if url_col is None:
        # Fall back to first column
        url_col = headers[0] if headers else None
    if name_col is None and len(headers) >= 2:
        # Only use the second column if it looks like a filename column
        candidate = headers[1]
        candidate_lower = candidate.lower()
        skip_words = {"desc", "description", "title", "caption", "alt", "tag", "category", "price", "sku"}
        if not any(w in candidate_lower for w in skip_words):
            name_col = candidate

    rows = []
    for raw in raw_rows:
        url_val = str(raw.get(url_col, "") or "").strip() if url_col else ""
        name_val = str(raw.get(name_col, "") or "").strip() if name_col else ""
        if url_val:
            rows.append({"original_url": url_val, "new_filename": name_val})

    return rows, (url_col or ""), (name_col or "")


def _find_column(headers: list[str], candidates: list[str]) -> str | None:
    """Return the first header that matches any candidate (case-insensitive)."""
    lower_headers = {h.lower(): h for h in headers}
    for c in candidates:
        if c.lower() in lower_headers:
            return lower_headers[c.lower()]
    return None


def _resolve_filename(new_filename: str, original_url: str, content_type: str) -> str:
    """Return a sanitized final filename, inferring extension if needed."""
    if new_filename:
        base, ext = os.path.splitext(new_filename)
        if not ext:
            ext = _ext_from_content_type(content_type) or _ext_from_url(original_url) or ".jpg"
        return _sanitize(base) + ext.lower()
    else:
        # Use filename from original URL
        path = urlparse(original_url).path
        basename = os.path.basename(path) or "image"
        base, ext = os.path.splitext(basename)
        if not ext:
            ext = _ext_from_content_type(content_type) or ".jpg"
        return _sanitize(base) + ext.lower()


def _sanitize(name: str) -> str:
    """Lowercase, replace spaces and special chars with hyphens, collapse multiples."""
    name = name.lower().strip()
    name = re.sub(r"[^\w\-.]", "-", name)
    name = re.sub(r"-{2,}", "-", name)
    return name.strip("-")


def _ext_from_content_type(content_type: str) -> str:
    ext = mimetypes.guess_extension(content_type)
    # mimetypes can return .jpe for image/jpeg — normalise
    aliases = {".jpe": ".jpg", ".jpeg": ".jpg"}
    return aliases.get(ext, ext) if ext else ""


def _ext_from_url(url: str) -> str:
    path = urlparse(url).path
    _, ext = os.path.splitext(path)
    return ext.lower() if ext else ""


def _build_output_excel(results: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["original_url", "new_filename", "new_url", "status", "message"])

    for r in results:
        ws.append([
            r["original_url"],
            r["new_filename"],
            r["new_url"],
            r["status"],
            r.get("message", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    app.run(debug=True, port=5001)
