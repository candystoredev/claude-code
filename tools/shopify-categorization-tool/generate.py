#!/usr/bin/env python3
"""
Shopify product categorization tool using Claude API.

Reads a Matrixify XLSX export of products and assigns Custom Collections
to each product based on its title, description, and vendor.

Usage:
    # Test on first 20 products
    python generate.py --limit 20

    # Process all products
    python generate.py

    # Resume from previous run
    python generate.py --resume

    # Custom input/output files
    python generate.py --input input/my_products.xlsx --output output/results.csv
"""

import argparse
import csv
import json
import os
import re
import sys
import time

import anthropic
import openpyxl

import config
from prompt_template import build_system_prompt, build_user_prompt


def load_collections() -> dict[str, str]:
    """Load the collections handle->title mapping from JSON."""
    with open(config.COLLECTIONS_FILE, encoding="utf-8") as f:
        return json.load(f)


def load_products_xlsx(filepath: str) -> list[dict]:
    """Load products from a Matrixify XLSX export.

    Expects columns: Handle, Title, Body HTML, Vendor
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)

    # Try common sheet names
    sheet_name = None
    for name in wb.sheetnames:
        if "product" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Parse headers (case-insensitive matching)
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    header_map = {}
    for i, h in enumerate(raw_headers):
        hl = h.lower()
        if hl == "handle":
            header_map["handle"] = i
        elif hl == "title":
            header_map["title"] = i
        elif hl in ("body html", "body (html)", "body_html"):
            header_map["body_html"] = i
        elif hl == "vendor":
            header_map["vendor"] = i

    if "handle" not in header_map:
        print(f"Error: Could not find 'Handle' column in {filepath}")
        print(f"Found headers: {raw_headers}")
        sys.exit(1)

    products = []
    for row in rows[1:]:
        handle = row[header_map["handle"]] if header_map.get("handle") is not None else ""
        if not handle:
            continue
        product = {
            "handle": str(handle).strip(),
            "title": str(row[header_map.get("title", 0)] or "").strip(),
            "body_html": str(row[header_map.get("body_html", 0)] or "").strip(),
            "vendor": str(row[header_map.get("vendor", 0)] or "").strip(),
        }
        products.append(product)

    wb.close()
    return products


def load_completed_handles(filepath: str) -> set[str]:
    """Load already-processed handles from the output CSV."""
    completed = set()
    if not os.path.exists(filepath):
        return completed
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            handle = row.get("Handle", "").strip()
            if handle and row.get("Custom Collections", "").strip():
                completed.add(handle)
    return completed


def categorize_batch(
    client: anthropic.Anthropic,
    system_prompt: str,
    products: list[dict],
    valid_handles: set[str],
    usage_stats: dict,
) -> dict[str, list[str]]:
    """Call Claude API to categorize a batch of products.

    Returns dict mapping product handle -> list of collection handles.
    Uses prompt caching to avoid re-sending the 258-collection system prompt each time.
    """
    user_prompt = build_user_prompt(products)

    message = client.messages.create(
        model=config.API_MODEL,
        max_tokens=config.API_MAX_TOKENS,
        # System prompt as content block with cache_control for prompt caching.
        # The collections list (~2K tokens) gets cached and reused across all batches.
        system=[
            {
                "type": "text",
                "text": system_prompt,
                "cache_control": {"type": "ephemeral"},
            }
        ],
        messages=[{"role": "user", "content": user_prompt}],
    )

    # Track token usage for cost reporting
    usage = message.usage
    usage_stats["input_tokens"] += usage.input_tokens
    usage_stats["output_tokens"] += usage.output_tokens
    usage_stats["cache_creation_tokens"] += getattr(usage, "cache_creation_input_tokens", 0)
    usage_stats["cache_read_tokens"] += getattr(usage, "cache_read_input_tokens", 0)

    response_text = message.content[0].text.strip()

    # Strip markdown code fences if present
    response_text = re.sub(r"^```(?:json)?\s*\n?", "", response_text)
    response_text = re.sub(r"\n?```\s*$", "", response_text)

    result = json.loads(response_text)

    # Validate: only keep handles that exist in our collections
    validated = {}
    for product_handle, collections in result.items():
        valid_collections = [c for c in collections if c in valid_handles]
        validated[product_handle] = valid_collections

    return validated


def init_output_csv(filepath: str):
    """Create the output CSV with headers if it doesn't exist."""
    if not os.path.exists(filepath):
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Handle", "Title", "Vendor", "Custom Collections"])


def append_results(filepath: str, results: list[dict]):
    """Append categorized products to the output CSV."""
    with open(filepath, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for r in results:
            writer.writerow([
                r["handle"],
                r["title"],
                r["vendor"],
                r["collections"],
            ])


def main():
    parser = argparse.ArgumentParser(
        description="Categorize Shopify products into Custom Collections using Claude API"
    )
    parser.add_argument(
        "--input", default=None,
        help="Input XLSX path (default: input/products.xlsx)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output CSV path (default: output/products_categorized.csv)"
    )
    parser.add_argument(
        "--limit", type=int, default=None,
        help="Process only the first N products (for testing)"
    )
    parser.add_argument(
        "--batch-size", type=int, default=config.BATCH_SIZE,
        help=f"Products per API call (default: {config.BATCH_SIZE})"
    )
    parser.add_argument(
        "--resume", action="store_true",
        help="Resume from previous run, skipping already-processed handles"
    )
    args = parser.parse_args()

    input_path = args.input or os.path.join(config.INPUT_DIR, config.DEFAULT_INPUT_FILE)
    output_path = args.output or os.path.join(config.OUTPUT_DIR, config.DEFAULT_OUTPUT_FILE)

    # Validate API key
    api_key = config.ANTHROPIC_API_KEY
    if not api_key:
        env_path = os.path.join(config.TOOL_DIR, ".env")
        if os.path.exists(env_path):
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("ANTHROPIC_API_KEY="):
                        api_key = line.split("=", 1)[1].strip().strip('"').strip("'")
                        break
    if not api_key:
        print("Error: ANTHROPIC_API_KEY not set.")
        print("Set it via environment variable or create a .env file in this directory:")
        print("  echo 'ANTHROPIC_API_KEY=sk-ant-...' > .env")
        sys.exit(1)

    # Load collections
    if not os.path.exists(config.COLLECTIONS_FILE):
        print(f"Error: Collections file not found: {config.COLLECTIONS_FILE}")
        sys.exit(1)

    collections = load_collections()
    valid_handles = set(collections.keys()) - {"show-all-products", "frontpage"}
    print(f"Loaded {len(valid_handles)} collection handles")

    # Build system prompt (done once, reused for all batches)
    system_prompt = build_system_prompt(collections)

    # Load products
    if not os.path.exists(input_path):
        print(f"Error: Input file not found: {input_path}")
        print(f"Place your Matrixify XLSX export in the input/ directory as '{config.DEFAULT_INPUT_FILE}'")
        sys.exit(1)

    products = load_products_xlsx(input_path)
    print(f"Loaded {len(products)} products from {input_path}")

    if args.limit:
        products = products[: args.limit]
        print(f"Limited to first {args.limit} products (test mode)")

    # Handle resume
    completed_handles = set()
    if args.resume:
        completed_handles = load_completed_handles(output_path)
        print(f"Resuming: {len(completed_handles)} products already processed")
    else:
        init_output_csv(output_path)

    # Filter out already-processed products
    remaining = [p for p in products if p["handle"] not in completed_handles]
    print(f"Products to process: {len(remaining)}")

    if not remaining:
        print("Nothing to process.")
        return

    client = anthropic.Anthropic(api_key=api_key)

    # Process in batches
    batch_size = args.batch_size
    total_batches = (len(remaining) + batch_size - 1) // batch_size
    processed = 0
    errors = 0
    buffer = []
    usage_stats = {
        "input_tokens": 0,
        "output_tokens": 0,
        "cache_creation_tokens": 0,
        "cache_read_tokens": 0,
    }

    print(f"Using model: {config.API_MODEL} with prompt caching enabled")
    print(f"Batch size: {batch_size} products/call ({total_batches} API calls total)\n")

    for batch_num in range(total_batches):
        start = batch_num * batch_size
        end = min(start + batch_size, len(remaining))
        batch = remaining[start:end]

        batch_handles = [p["handle"] for p in batch]
        print(
            f"[Batch {batch_num + 1}/{total_batches}] "
            f"Processing {len(batch)} products: {batch_handles[0]}...{batch_handles[-1]}",
            end=" ",
            flush=True,
        )

        try:
            result = categorize_batch(client, system_prompt, batch, valid_handles, usage_stats)

            for p in batch:
                collections_list = result.get(p["handle"], [])
                buffer.append({
                    "handle": p["handle"],
                    "title": p["title"],
                    "vendor": p["vendor"],
                    "collections": ", ".join(collections_list),
                })
                processed += 1

            avg_collections = sum(len(v) for v in result.values()) / max(len(result), 1)
            print(f"OK (avg {avg_collections:.1f} collections/product)")

        except json.JSONDecodeError as e:
            print(f"JSON PARSE ERROR: {e}")
            for p in batch:
                buffer.append({
                    "handle": p["handle"],
                    "title": p["title"],
                    "vendor": p["vendor"],
                    "collections": "",
                })
            errors += len(batch)

        except anthropic.APIError as e:
            print(f"API ERROR: {e}")
            for p in batch:
                buffer.append({
                    "handle": p["handle"],
                    "title": p["title"],
                    "vendor": p["vendor"],
                    "collections": "",
                })
            errors += len(batch)

        except Exception as e:
            print(f"ERROR: {e}")
            for p in batch:
                buffer.append({
                    "handle": p["handle"],
                    "title": p["title"],
                    "vendor": p["vendor"],
                    "collections": "",
                })
            errors += len(batch)

        # Checkpoint
        if len(buffer) >= config.SAVE_EVERY_N * batch_size:
            append_results(output_path, buffer)
            buffer = []
            print(f"  -- Checkpoint saved. Processed: {processed}, Errors: {errors}")

        # Rate limiting
        time.sleep(1.0 / config.REQUESTS_PER_SECOND)

    # Flush remaining buffer
    if buffer:
        append_results(output_path, buffer)

    print("\n--- Done ---")
    print(f"Processed: {processed}")
    print(f"Errors: {errors}")
    print(f"Output: {output_path}")

    # Cost report
    print("\n--- Token Usage ---")
    print(f"Input tokens:          {usage_stats['input_tokens']:,}")
    print(f"Output tokens:         {usage_stats['output_tokens']:,}")
    print(f"Cache creation tokens: {usage_stats['cache_creation_tokens']:,}")
    print(f"Cache read tokens:     {usage_stats['cache_read_tokens']:,}")
    if usage_stats["cache_read_tokens"] > 0:
        total_input = usage_stats["input_tokens"] + usage_stats["cache_read_tokens"]
        cache_hit_pct = usage_stats["cache_read_tokens"] / total_input * 100
        print(f"Cache hit rate:        {cache_hit_pct:.1f}%")
        saved = usage_stats["cache_read_tokens"] * 0.9  # Cache reads are 90% cheaper
        print(f"Tokens saved by cache: ~{saved:,.0f} (90% discount on cached input)")


if __name__ == "__main__":
    main()
