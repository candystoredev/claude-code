import os
import io
import json
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import load_workbook, Workbook
from anthropic import Anthropic, AuthenticationError, APIError
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB limit

api_key = os.environ.get("ANTHROPIC_API_KEY", "")
client = Anthropic(api_key=api_key) if api_key else None

HANDLE_RULES_PROMPT = """You are a Shopify handle generator. Given product names, generate URL-friendly handles following these rules EXACTLY:

1. All lowercase, single hyphens only (no spaces, capitals, underscores).
2. Remove apostrophes (1950's → 1950s).
3. Omit "&" entirely (Mike & Ike → mike-ike).
4. Omit short articles/conjunctions: "and", "with", "of", "the", "for", "in", "a", "an", etc.
5. Shorten aggressively: keep only core keywords (brand/flavor/product); drop redundant words ("candy", "gummy" when obvious from context).
6. Drop lesser-known/manufacturer brands entirely (Nassau Candy, Clever Candy, Madelaine, etc.). Only keep strong consumer brands (mike-ike, skittles, hersheys, lindt, etc.).
7. Unit size/oz/ct/lb: Leave out completely unless needed to distinguish exact duplicate titles (then add minimal distinguisher like -12ct). No decimals in handles ever.
8. Packaging words: Minimize/omit (peg-bag, theater-box, tubs, pouches, tins, etc.) unless needed for uniqueness.
9. Flavor/variant order: Match title rules (sour first if leading title, brand first otherwise).
10. Goal: Short, clean, readable URLs (e.g., mike-ike-mega-mix, sour-patch-watermelon).

You MUST return ONLY valid JSON — an array of objects with "product_name" and "handle" fields. No markdown, no explanation, no code fences. Just the raw JSON array.

Track all handles in this batch and ensure uniqueness. If two products would produce the same handle, add a minimal distinguisher to the second one."""


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate_handles():
    if not client:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set. Add it to your .env file and restart the server."}), 500

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid request body"}), 400

    product_names = data.get("product_names", [])
    existing_handles = data.get("existing_handles", [])

    if not product_names:
        return jsonify({"error": "No product names provided"}), 400

    try:
        user_message = "Generate Shopify handles for these products:\n\n"
        for j, name in enumerate(product_names, 1):
            user_message += f"{j}. {name}\n"

        # Only send the most recent handles to avoid prompt bloat/timeouts
        # on large files. 100 recent handles is enough for dedup context.
        if existing_handles:
            recent_handles = existing_handles[-100:]
            user_message += (
                "\n\nAlready-used handles (must not duplicate): "
                + ", ".join(recent_handles)
            )

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": user_message}],
            system=HANDLE_RULES_PROMPT,
        )

        response_text = message.content[0].text.strip()

        # Strip markdown code fences if present
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            lines = [l for l in lines[1:] if l.strip() != "```"]
            response_text = "\n".join(lines)

        try:
            results = json.loads(response_text)
        except json.JSONDecodeError:
            return jsonify({"error": "Failed to parse AI response", "raw": response_text}), 500

        return jsonify({"results": results})

    except AuthenticationError:
        return jsonify({"error": "Invalid API key. Check your ANTHROPIC_API_KEY in the .env file."}), 401
    except APIError as e:
        return jsonify({"error": f"Anthropic API error: {str(e)}"}), 502
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500


PRODUCT_HEADERS = {"product", "product name", "product_name", "title", "name", "item", "item name", "product title", "product-name"}
SKIP_HEADERS = {"sku", "id", "upc", "barcode", "code", "product code", "item number", "item_number", "item #", "sku #"}


def _parse_excel(file):
    wb = load_workbook(file)
    ws = wb.active
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or "").lower().strip()
        headers[col] = val

    product_col = None
    detected_header = None

    # First pass: look for explicit product name headers
    for col, header in headers.items():
        if header in PRODUCT_HEADERS:
            product_col = col
            detected_header = header
            break

    # Second pass: pick the first text-heavy column that isn't a known skip column
    if product_col is None:
        for col in range(1, ws.max_column + 1):
            header = headers.get(col, "")
            if header in SKIP_HEADERS:
                continue
            # Check if column values look like product names (longer text, not codes)
            sample_vals = []
            for r in range(2, min(ws.max_row + 1, 7)):
                v = ws.cell(row=r, column=col).value
                if v:
                    sample_vals.append(str(v).strip())
            if sample_vals:
                avg_len = sum(len(v) for v in sample_vals) / len(sample_vals)
                if avg_len > 10:  # Product names are typically longer than SKUs/codes
                    product_col = col
                    detected_header = header or f"Column {col}"
                    break

    # Final fallback: first non-skip column
    if product_col is None:
        for col in range(1, ws.max_column + 1):
            header = headers.get(col, "")
            if header not in SKIP_HEADERS:
                product_col = col
                detected_header = header or f"Column {col}"
                break
        if product_col is None:
            product_col = 1
            detected_header = headers.get(1, "Column 1")

    product_names = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=product_col).value
        if val and str(val).strip():
            product_names.append(str(val).strip())

    return product_names, detected_header


def _parse_csv(file):
    import csv
    content = file.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    if not rows:
        return [], "empty"

    headers = [h.lower().strip() for h in rows[0]]

    product_col = None
    detected_header = None

    for i, header in enumerate(headers):
        if header in PRODUCT_HEADERS:
            product_col = i
            detected_header = header
            break

    if product_col is None:
        for i, header in enumerate(headers):
            if header in SKIP_HEADERS:
                continue
            sample_vals = [rows[r][i] for r in range(1, min(len(rows), 6)) if i < len(rows[r]) and rows[r][i].strip()]
            if sample_vals:
                avg_len = sum(len(v) for v in sample_vals) / len(sample_vals)
                if avg_len > 10:
                    product_col = i
                    detected_header = header or f"Column {i + 1}"
                    break

    if product_col is None:
        for i, header in enumerate(headers):
            if header not in SKIP_HEADERS:
                product_col = i
                detected_header = header or f"Column {i + 1}"
                break
        if product_col is None:
            product_col = 0
            detected_header = headers[0] if headers else "Column 1"

    product_names = []
    for row in rows[1:]:
        if product_col < len(row) and row[product_col].strip():
            product_names.append(row[product_col].strip())

    return product_names, detected_header


@app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    if filename.endswith((".xlsx", ".xls")):
        product_names, detected_header = _parse_excel(file)
    elif filename.endswith(".csv"):
        product_names, detected_header = _parse_csv(file)
    else:
        return jsonify({"error": "Please upload an Excel (.xlsx) or CSV (.csv) file"}), 400

    if not product_names:
        return jsonify({"error": "No product names found. Make sure your file has a column with product names (see format guide below)."}), 400

    return jsonify({
        "product_names": product_names,
        "detected_column": detected_header,
    })


@app.route("/download", methods=["POST"])
def download_file():
    data = request.get_json()
    results = data.get("results", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Shopify Handles"
    ws.append(["Product Name", "Handle"])

    for r in results:
        ws.append([r["product_name"], r["handle"]])

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="shopify_handles.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.xml",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(debug=debug, host="0.0.0.0", port=port)
