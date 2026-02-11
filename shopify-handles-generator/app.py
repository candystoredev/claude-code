import os
import io
import json
import tempfile
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import load_workbook, Workbook
from anthropic import Anthropic, AuthenticationError, APIError
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB limit

api_key = os.environ.get("ANTHROPIC_API_KEY", "")
client = Anthropic(api_key=api_key) if api_key else None

HANDLE_RULES_PROMPT = """You are a Shopify handle generator. Given product names, generate URL-friendly handles following these rules EXACTLY:

1. All lowercase, single hyphens only (no spaces, capitals, underscores).
2. Remove apostrophes (1950's → 1950s).
3. Omit "&" entirely (Mike & Ike → mike-ike).
4. Omit short articles/conjunctions: "and", "with", "of", "the", "for", "in", "a", "an", etc.
5. Shorten aggressively: keep only core keywords (brand/flavor/product); drop redundant words ("candy", "gummy" when obvious from context).
6. Drop lesser-known/manufacturer brands entirely (Nassau Candy, Clever Candy, Madelaine, etc.). Only keep strong consumer brands (mike-ike, skittles, hersheys, lindt, etc.).
7. Unit size/oz/ct/lb: Leave out completely unless needed to distinguish exact duplicate titles (then add minimal distinguisher like -12ct). No decimals in handles ever.
8. Packaging words: Minimize/omit (peg-bag, theater-box, tubs, pouches, tins, etc.) unless needed for uniqueness.
9. Flavor/variant order: Match title rules (sour first if leading title, brand first otherwise).
10. Goal: Short, clean, readable URLs (e.g., mike-ike-mega-mix, sour-patch-watermelon).

You MUST return ONLY valid JSON — an array of objects with "product_name" and "handle" fields. No markdown, no explanation, no code fences. Just the raw JSON array.

Track all handles in this batch and ensure uniqueness. If two products would produce the same handle, add a minimal distinguisher to the second one."""


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate_handles():
    if not client:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set. Add it to your .env file and restart the server."}), 500

    data = request.get_json()
    product_names = data.get("product_names", [])
    existing_handles = data.get("existing_handles", [])

    if not product_names:
        return jsonify({"error": "No product names provided"}), 400

    try:
        user_message = "Generate Shopify handles for these products:\n\n"
        for j, name in enumerate(product_names, 1):
            user_message += f"{j}. {name}\n"

        if existing_handles:
            user_message += (
                "\n\nAlready-used handles (must not duplicate): "
                + ", ".join(existing_handles)
            )

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": user_message}],
            system=HANDLE_RULES_PROMPT,
        )

        response_text = message.content[0].text.strip()

        # Strip markdown code fences if present
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            lines = [l for l in lines[1:] if l.strip() != "```"]
            response_text = "\n".join(lines)

        try:
            results = json.loads(response_text)
        except json.JSONDecodeError:
            return jsonify({"error": "Failed to parse AI response", "raw": response_text}), 500

        return jsonify({"results": results})

    except AuthenticationError:
        return jsonify({"error": "Invalid API key. Check your ANTHROPIC_API_KEY in the .env file."}), 401
    except APIError as e:
        return jsonify({"error": f"Anthropic API error: {str(e)}"}), 502
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel file (.xlsx)"}), 400

    wb = load_workbook(file)
    ws = wb.active

    # Find the column with product names (look for common headers)
    product_col = None
    header_row = 1
    product_names = []

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value or "").lower().strip()
        if header in ("product", "product name", "product_name", "title", "name", "item", "item name"):
            product_col = col
            break

    if product_col is None:
        # Default to first column
        product_col = 1
        # Check if first row looks like a header
        first_val = str(ws.cell(row=1, column=1).value or "")
        if first_val and not first_val[0].isdigit():
            header_row = 1
        else:
            header_row = 0

    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=product_col).value
        if val and str(val).strip():
            product_names.append(str(val).strip())

    return jsonify({
        "product_names": product_names,
        "product_col": product_col,
        "header_row": header_row,
    })


@app.route("/download", methods=["POST"])
def download_file():
    data = request.get_json()
    results = data.get("results", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Shopify Handles"
    ws.append(["Product Name", "Handle"])

    for r in results:
        ws.append([r["product_name"], r["handle"]])

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="shopify_handles.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.xml",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)
